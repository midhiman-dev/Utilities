#!/usr/bin/env python3
r"""Backlog DOCX to Excel converter.

Recognition:
- Epic headings: Heading 2 (or higher) whose text matches ^EP\d{2}\s*[–-]\s*
- User stories table: the first table after an epic meta table that has columns
    resembling story_id/summary/role/acceptance criteria (case-insensitive)
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert backlog DOCX to structured Excel.")
    parser.add_argument("--input", required=True, help="Path to backlog .docx")
    parser.add_argument("--output", required=False, help="Path to output .xlsx")
    parser.add_argument("--verbose", action="store_true", help="Print progress logs")
    parser.add_argument("--emit-csv", action="store_true", help="Write Jira CSV files")
    parser.add_argument("--csv-dir", required=False, help="Directory to place CSVs")
    parser.add_argument("--csv-encoding", default="utf-8", help="CSV encoding (default utf-8)")
    parser.add_argument("--csv-delimiter", default=",", help="CSV delimiter (default ,)")
    parser.add_argument(
        "--csv-quote",
        default="minimal",
        choices=["minimal", "all"],
        help="CSV quote mode: minimal|all",
    )
    parser.add_argument("--csv-overwrite", action="store_true", help="Overwrite existing CSVs")
    return parser.parse_args(argv)
from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph

EPIC_HEADING_RE = re.compile(r"^(EP\d{2})\s*[–-]\s*(.+)$", re.IGNORECASE)

EPIC_INDEX_COLUMNS = [
    "epic_id",
    "epic_name",
    "objective_summary",
    "priority",
    "primary_screens",
    "key_tables",
]

EPIC_COLUMNS = [
    "epic_id",
    "epic_name",
    "objective_summary",
    "priority",
    "primary_screens",
    "key_tables",
    "components",
    "notes",
]

STORY_COLUMNS = [
    "epic_id",
    "story_id",
    "summary",
    "primary_role",
    "ui_screens",
    "key_data_fields",
    "acceptance_criteria",
    "priority",
    "estimated_effort_hrs",
]

DOC_CONTROL_COLUMNS = ["document_name", "purpose", "version", "date", "notes"]

EPIC_CSV_COLUMNS = ["Issue Type", "Summary", "Epic Name", "Description", "Priority", "Labels"]

STORY_CSV_COLUMNS = [
    "Issue Type",
    "Summary",
    "Description",
    "Priority",
    "Labels",
    "Epic Link",
    "Epic Name (for lookup)",
    "Primary Role",
    "UI Screens",
    "Key Data Fields",
    "Acceptance Criteria",
    "Estimated Effort (hrs)",
]


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip()).lower()


def _cell_text(cell) -> str:
    parts: List[str] = []
    for para in cell.paragraphs:
        line = para.text.rstrip()
        if line:
            parts.append(line)
    return "\n".join(parts)


def _table_to_rows(table: Table) -> List[List[str]]:
    rows: List[List[str]] = []
    for row in table.rows:
        rows.append([_cell_text(c) for c in row.cells])
    return rows


def _find_column_map(headers: List[str], mapping: Dict[str, List[str]]) -> Dict[str, int]:
    normalized = [_norm(h) for h in headers]
    col_map: Dict[str, int] = {}
    for canonical, options in mapping.items():
        for i, h in enumerate(normalized):
            if h in options:
                col_map[canonical] = i
                break
    return col_map


def _is_epic_index_table(table: Table) -> bool:
    rows = _table_to_rows(table)
    if not rows:
        return False
    headers = rows[0]
    mapping = {
        "epic_id": ["epic id", "epic"],
        "epic_name": ["epic name", "name"],
        "objective_summary": ["objective (summary)", "objective", "summary"],
        "priority": ["priority"],
        "primary_screens": ["primary screens", "screens"],
        "key_tables": ["key tables", "tables", "key data tables"],
    }
    col_map = _find_column_map(headers, mapping)
    return len(col_map) >= 4 and "epic_id" in col_map and "epic_name" in col_map


def _is_story_table(table: Table) -> bool:
    rows = _table_to_rows(table)
    if not rows:
        return False
    headers = rows[0]
    mapping = {
        "story_id": ["story id", "user story id", "id"],
        "summary": ["summary", "user story", "story"],
        "primary_role": ["primary role", "role"],
        "ui_screens": ["ui screens", "screens", "primary screens"],
        "key_data_fields": ["key data fields", "key data", "data fields"],
        "acceptance_criteria": ["acceptance criteria", "acceptance criteria (high-level)", "acceptance"],
        "priority": ["priority"],
    }
    col_map = _find_column_map(headers, mapping)
    return "story_id" in col_map and "summary" in col_map


def _parse_doc_control(table: Table) -> Dict[str, str]:
    rows = _table_to_rows(table)
    result: Dict[str, str] = {}
    for row in rows:
        if len(row) < 2:
            continue
        key = _norm(row[0])
        val = row[1].strip()
        if not key:
            continue
        if "document" in key and "name" in key:
            result["document_name"] = val
        elif "purpose" in key:
            result["purpose"] = val
        elif "version" in key:
            result["version"] = val
        elif "date" in key:
            result["date"] = val
        elif "notes" in key:
            result["notes"] = val
    return result


def _table_as_kv(table: Table) -> Dict[str, str]:
    rows = _table_to_rows(table)
    data: Dict[str, str] = {}
    for row in rows:
        if len(row) < 2:
            continue
        key = _norm(row[0])
        val = row[1].strip()
        if key:
            data[key] = val
    return data


def _iter_block_items(doc: Document):
    parent = doc.element.body
    for child in parent.iterchildren():
        if child.tag.endswith("}p"):
            yield Paragraph(child, doc)
        elif child.tag.endswith("}tbl"):
            yield Table(child, doc)


def parse_docx(path: str) -> Dict[str, object]:
    doc = Document(path)

    epics: List[Dict[str, str]] = []
    stories: List[Dict[str, str]] = []
    epic_index: List[Dict[str, str]] = []
    questions: List[Dict[str, str]] = []
    doc_control: Dict[str, str] = {}

    current_epic: Optional[Dict[str, str]] = None
    seen_epic_index = False
    in_questions = False

    for block in _iter_block_items(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue

            style = (block.style.name or "").lower()
            m = EPIC_HEADING_RE.match(text)
            if m and "heading" in style:
                epic_id = m.group(1).upper()
                epic_name = m.group(2).strip()
                current_epic = {
                    "epic_id": epic_id,
                    "epic_name": epic_name,
                    "objective_summary": "",
                    "priority": "",
                    "primary_screens": "",
                    "key_tables": "",
                    "components": "",
                    "notes": "",
                }
                epics.append(current_epic)
                in_questions = False
                continue

            if re.search(r"questions\s*/\s*clarifications|questions|clarifications", text, re.IGNORECASE):
                in_questions = True
                continue

            if in_questions:
                questions.append({"item_no": str(len(questions) + 1), "question_or_note": text})
            continue

        # Table handling
        table = block
        if not seen_epic_index and _is_epic_index_table(table):
            rows = _table_to_rows(table)
            headers = rows[0]
            mapping = {
                "epic_id": ["epic id", "epic"],
                "epic_name": ["epic name", "name"],
                "objective_summary": ["objective (summary)", "objective", "summary"],
                "priority": ["priority"],
                "primary_screens": ["primary screens", "screens"],
                "key_tables": ["key tables", "tables", "key data tables"],
            }
            col_map = _find_column_map(headers, mapping)
            for row in rows[1:]:
                if not any(cell.strip() for cell in row):
                    continue
                item = {k: (row[idx].strip() if idx < len(row) else "") for k, idx in col_map.items()}
                epic_index.append(item)
            seen_epic_index = True
            continue

        # Document Control (2-column table near top)
        if not doc_control and len(table.columns) == 2:
            dc = _parse_doc_control(table)
            if dc:
                doc_control = dc
                continue

        if current_epic is None:
            continue

        # Epic meta table (key-value)
        if len(table.columns) == 2 and not _is_story_table(table):
            kv = _table_as_kv(table)
            if not kv:
                continue
            key_map = {
                "priority": ["priority"],
                "components": ["components", "component"],
                "primary_screens": ["primary screens", "screens"],
                "key_tables": ["key data tables", "key tables", "tables"],
                "notes": ["notes", "note"],
            }
            for target, options in key_map.items():
                for key in options:
                    if key in kv and kv[key]:
                        current_epic[target] = kv[key]
                        break
            continue

        # User stories table
        if _is_story_table(table):
            rows = _table_to_rows(table)
            headers = rows[0]
            mapping = {
                "story_id": ["story id", "user story id", "id"],
                "summary": ["summary", "user story", "story"],
                "primary_role": ["primary role", "role"],
                "ui_screens": ["ui screens", "screens", "primary screens"],
                "key_data_fields": ["key data fields", "key data", "data fields"],
                "acceptance_criteria": ["acceptance criteria", "acceptance criteria (high-level)", "acceptance"],
                "priority": ["priority"],
            }
            col_map = _find_column_map(headers, mapping)
            for row in rows[1:]:
                if not any(cell.strip() for cell in row):
                    continue
                story = {
                    "epic_id": current_epic["epic_id"],
                    "story_id": row[col_map.get("story_id", -1)].strip() if col_map.get("story_id", -1) >= 0 else "",
                    "summary": row[col_map.get("summary", -1)].strip() if col_map.get("summary", -1) >= 0 else "",
                    "primary_role": row[col_map.get("primary_role", -1)].strip() if col_map.get("primary_role", -1) >= 0 else "",
                    "ui_screens": row[col_map.get("ui_screens", -1)].strip() if col_map.get("ui_screens", -1) >= 0 else "",
                    "key_data_fields": row[col_map.get("key_data_fields", -1)].strip() if col_map.get("key_data_fields", -1) >= 0 else "",
                    "acceptance_criteria": row[col_map.get("acceptance_criteria", -1)].strip() if col_map.get("acceptance_criteria", -1) >= 0 else "",
                    "priority": row[col_map.get("priority", -1)].strip() if col_map.get("priority", -1) >= 0 else "",
                    "estimated_effort_hrs": "",
                }
                stories.append(story)
            continue

    # Merge epic index into epics when possible
    if epic_index:
        idx_map = {e.get("epic_id", "").upper(): e for e in epic_index if e.get("epic_id")}
        for epic in epics:
            idx = idx_map.get(epic["epic_id"].upper())
            if not idx:
                continue
            epic["objective_summary"] = idx.get("objective_summary", "") or epic["objective_summary"]
            epic["priority"] = idx.get("priority", "") or epic["priority"]
            epic["primary_screens"] = idx.get("primary_screens", "") or epic["primary_screens"]
            epic["key_tables"] = idx.get("key_tables", "") or epic["key_tables"]

    return {
        "epics": epics,
        "stories": stories,
        "epic_index": epic_index,
        "questions": questions,
        "doc_control": doc_control,
    }


def _autosize_columns(ws, df: pd.DataFrame, wrap_cols: List[str], max_width: int = 60):
    from openpyxl.styles import Alignment, Font

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), 10)
        for v in df[col_name].astype(str).fillna(""):
            max_len = max(max_len, len(v))
        width = min(max_len + 2, max_width)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        if col_name in wrap_cols:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(width, 18)

    for col_name in wrap_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).alignment = wrap


def write_excel(structured: Dict[str, object], out_path: str) -> None:
    epics = pd.DataFrame(structured["epics"], columns=EPIC_COLUMNS)
    stories = pd.DataFrame(structured["stories"], columns=STORY_COLUMNS)

    sheet_frames: Dict[str, pd.DataFrame] = {
        "Epics": epics,
        "User_Stories": stories,
    }

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        epics.to_excel(writer, sheet_name="Epics", index=False)
        stories.to_excel(writer, sheet_name="User_Stories", index=False)

        if structured.get("epic_index"):
            df = pd.DataFrame(structured["epic_index"], columns=EPIC_INDEX_COLUMNS)
            df.to_excel(writer, sheet_name="Epic_Index", index=False)
            sheet_frames["Epic_Index"] = df

        if structured.get("questions"):
            df = pd.DataFrame(structured["questions"], columns=["item_no", "question_or_note"])
            df.to_excel(writer, sheet_name="Questions_Clarifications", index=False)
            sheet_frames["Questions_Clarifications"] = df

        if structured.get("doc_control"):
            df = pd.DataFrame([structured["doc_control"]], columns=DOC_CONTROL_COLUMNS)
            df.to_excel(writer, sheet_name="Document_Control", index=False)
            sheet_frames["Document_Control"] = df

        for sheet_name, df in sheet_frames.items():
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            wrap_cols = [
                "objective_summary",
                "key_data_fields",
                "acceptance_criteria",
                "notes",
            ]
            _autosize_columns(ws, df, wrap_cols)
            if sheet_name == "User_Stories" and "estimated_effort_hrs" in df.columns:
                col_idx = df.columns.get_loc("estimated_effort_hrs") + 1
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18


def _build_output_path(input_path: Path, output_arg: Optional[str]) -> Path:
    if output_arg:
        return Path(output_arg)
    return input_path.with_suffix(".xlsx")


def build_epics_csv_rows(epics: List[Dict[str, str]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for epic in epics:
        description = epic.get("objective_summary", "").strip()
        if not description:
            description = epic.get("notes", "").strip() or epic.get("components", "").strip()
        rows.append(
            {
                "Issue Type": "Epic",
                "Summary": epic.get("epic_name", ""),
                "Epic Name": epic.get("epic_name", ""),
                "Description": description,
                "Priority": epic.get("priority", ""),
                "Labels": "",
            }
        )
    return rows


def build_stories_csv_rows(
    stories: List[Dict[str, str]], epic_lookup: Dict[str, str]
) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for story in stories:
        epic_id = story.get("epic_id", "")
        epic_name = epic_lookup.get(epic_id, "")
        description = (
            f"Primary Role: {story.get('primary_role', '')}\n"
            f"UI Screens: {story.get('ui_screens', '')}\n"
            f"Key Data Fields: {story.get('key_data_fields', '')}\n"
            "Acceptance Criteria:\n"
            f"{story.get('acceptance_criteria', '')}"
        )
        rows.append(
            {
                "Issue Type": "Story",
                "Summary": story.get("summary", ""),
                "Description": description,
                "Priority": story.get("priority", ""),
                "Labels": "",
                "Epic Link": "",
                "Epic Name (for lookup)": epic_name,
                "Primary Role": story.get("primary_role", ""),
                "UI Screens": story.get("ui_screens", ""),
                "Key Data Fields": story.get("key_data_fields", ""),
                "Acceptance Criteria": story.get("acceptance_criteria", ""),
                "Estimated Effort (hrs)": "",
            }
        )
    return rows


def write_jira_csvs(
    structured: Dict[str, object],
    out_dir: Path,
    encoding: str,
    delimiter: str,
    quote_mode: str,
    overwrite: bool,
    verbose: bool,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    quote_map = {"minimal": csv.QUOTE_MINIMAL, "all": csv.QUOTE_ALL}
    quoting = quote_map.get(quote_mode, csv.QUOTE_MINIMAL)

    epics = structured.get("epics", [])
    stories = structured.get("stories", [])
    epic_lookup = {e.get("epic_id", ""): e.get("epic_name", "") for e in epics}

    epic_rows = build_epics_csv_rows(epics)
    story_rows = build_stories_csv_rows(stories, epic_lookup)

    paths = {
        "epics": out_dir / "epics.csv",
        "stories": out_dir / "stories.csv",
        "mapping": out_dir / "epic_key_mapping.csv",
    }
    if not overwrite:
        for p in paths.values():
            if p.exists():
                raise FileExistsError(f"CSV already exists: {p}")

    pd.DataFrame(epic_rows, columns=EPIC_CSV_COLUMNS).to_csv(
        paths["epics"],
        index=False,
        encoding=encoding,
        sep=delimiter,
        quoting=quoting,
    )
    pd.DataFrame(story_rows, columns=STORY_CSV_COLUMNS).to_csv(
        paths["stories"],
        index=False,
        encoding=encoding,
        sep=delimiter,
        quoting=quoting,
    )
    pd.DataFrame(
        [{"Epic Name": e.get("epic_name", ""), "Epic Key": ""} for e in epics],
        columns=["Epic Name", "Epic Key"],
    ).to_csv(
        paths["mapping"],
        index=False,
        encoding=encoding,
        sep=delimiter,
        quoting=quoting,
    )

    print(
        f"Wrote epics.csv ({len(epic_rows)} rows), stories.csv ({len(story_rows)} rows), epic_key_mapping.csv"
    )
    print(
        "NOTE: Import epics.csv first; then fill Epic Key in epic_key_mapping.csv and update stories.csv Epic Link before importing stories.csv."
    )
    if verbose:
        print(
            "Jira import guide: import epics.csv → copy Epic Keys into epic_key_mapping.csv → update stories.csv Epic Link → import stories.csv → validate backlog."
        )


def run(args: argparse.Namespace) -> int:

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        return 2

    if args.verbose:
        print(f"Reading: {input_path}")

    structured = parse_docx(str(input_path))
    epics = structured["epics"]
    stories = structured["stories"]
    exit_code = 0
    if not epics:
        print("Warning: no epics detected.", file=sys.stderr)
        exit_code = 3
    if not stories:
        print("Warning: no user stories detected.", file=sys.stderr)

    if args.verbose:
        print(f"Epics found: {len(epics)}")
        print(f"User Stories found: {len(stories)}")
        if structured.get("epic_index"):
            print(f"Epic Index rows: {len(structured['epic_index'])}")
        if structured.get("questions"):
            print(f"Questions/Clarifications items: {len(structured['questions'])}")
        print("User_Stories: added editable 'estimated_effort_hrs' column (blank by default)")

    for epic in epics:
        if not any(s["epic_id"] == epic["epic_id"] for s in stories):
            print(f"Warning: Epic {epic['epic_id']} has no User Stories table.", file=sys.stderr)

    for story in stories:
        if not story.get("story_id") or not story.get("summary"):
            print("Warning: Story missing Story ID or Summary.", file=sys.stderr)
            break

    out_path = _build_output_path(input_path, args.output)
    write_excel(structured, str(out_path))
    print(f"Output written: {out_path}")

    if args.emit_csv:
        csv_dir = Path(args.csv_dir) if args.csv_dir else out_path.parent
        try:
            write_jira_csvs(
                structured,
                csv_dir,
                args.csv_encoding,
                args.csv_delimiter,
                args.csv_quote,
                args.csv_overwrite,
                args.verbose,
            )
        except FileExistsError as exc:
            print(f"Error: {exc}", file=sys.stderr)
            exit_code = 4

        if any(not s.get("epic_id") for s in stories):
            print("Warning: Some stories missing epic_id; Epic Name lookup may be blank.", file=sys.stderr)

    return exit_code


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    return run(args)


if __name__ == "__main__":
    sys.exit(main())
